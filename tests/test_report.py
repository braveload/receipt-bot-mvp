from openpyxl import load_workbook

from app import report, storage
from app.models import ReceiptData


def test_excel_report_includes_amount_supply_and_vat_totals(tmp_path):
    storage.save_receipt(
        "report-user",
        ReceiptData(
            merchant="테스트상점",
            amount=11000,
            supply_amount=10000,
            vat_amount=1000,
            date="2026-07-25",
            doc_type="카드전표",
            category="소모품비",
            biz_or_personal="사업",
            confidence=0.95,
        ),
    )
    storage.save_receipt(
        "report-user",
        ReceiptData(
            merchant="개인지출",
            amount=5500,
            supply_amount=5000,
            vat_amount=500,
            date="2026-07-25",
            doc_type="카드전표",
            category="기타",
            biz_or_personal="개인",
            confidence=0.9,
        ),
    )

    output = report.export_to_excel("report-user", tmp_path / "report.xlsx")
    worksheet = load_workbook(output).active
    total_row = worksheet.max_row

    assert worksheet.cell(total_row, 1).value == "사업 경비 합계"
    assert worksheet.cell(total_row, 3).value == 11000
    assert worksheet.cell(total_row, 4).value == 10000
    assert worksheet.cell(total_row, 5).value == 1000
    assert worksheet.cell(total_row, 5).number_format == '#,##0"원"'
