"""월간 요약 텍스트 + 신고철 엑셀 내보내기."""
from __future__ import annotations

from pathlib import Path
from sqlite3 import Row

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import storage


def build_monthly_summary_data(kakao_user_id: str, yyyy_mm: str) -> dict:
    """월간 요약의 구조화된 데이터 (웹/앱 API용). 카톡 텍스트 요약도 이 값을 재사용한다."""
    rows = storage.get_receipts_for_month(kakao_user_id, yyyy_mm)

    biz_total = sum(r["amount"] for r in rows if r["biz_or_personal"] == "사업")
    personal_total = sum(r["amount"] for r in rows if r["biz_or_personal"] == "개인")
    by_category: dict[str, int] = {}
    for r in rows:
        by_category[r["category"]] = by_category.get(r["category"], 0) + r["amount"]
    top_categories = sorted(by_category.items(), key=lambda kv: kv[1], reverse=True)

    return {
        "month": yyyy_mm,
        "count": len(rows),
        "biz_total": biz_total,
        "personal_total": personal_total,
        "categories": [{"category": cat, "amount": amt} for cat, amt in top_categories],
    }


def build_monthly_summary_text(kakao_user_id: str, yyyy_mm: str) -> str:
    """카카오톡으로 보낼 월간 요약 메시지 (설계서 5.1 '월간 요약 리포트')."""
    data = build_monthly_summary_data(kakao_user_id, yyyy_mm)

    if data["count"] == 0:
        return f"{yyyy_mm}에는 아직 등록된 영수증이 없어요. 사진을 보내주시면 자동으로 정리해드릴게요!"

    top3 = data["categories"][:3]
    category_lines = "\n".join(f"  · {c['category']}: {c['amount']:,}원" for c in top3)

    return (
        f"[{yyyy_mm} 지출 요약]\n"
        f"총 {data['count']}건 처리했어요.\n\n"
        f"인정 가능 경비(사업): {data['biz_total']:,}원\n"
        f"개인 지출: {data['personal_total']:,}원\n\n"
        f"카테고리 TOP3\n{category_lines}\n\n"
        f"신고철에 세무사 전달용 파일이 필요하면 '신고파일'이라고 보내주세요."
    )


HEADER = [
    "일자", "상호명", "금액", "공급가액", "부가세", "증빙유형",
    "카테고리", "사업/개인", "사업자등록번호", "추출 신뢰도",
]


def export_to_excel(kakao_user_id: str, out_path: Path, yyyy_mm: str | None = None) -> Path:
    """설계서 5.1 '신고철 원클릭 내보내기' — 세무사 전달용 엑셀 생성."""
    rows: list[Row] = (
        storage.get_receipts_for_month(kakao_user_id, yyyy_mm)
        if yyyy_mm
        else storage.get_all_receipts(kakao_user_id)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "경비내역"

    ws.append(HEADER)
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    for col_idx, _ in enumerate(HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    biz_total = 0
    supply_total = 0
    vat_total = 0
    for r in rows:
        ws.append(
            [
                r["date"],
                r["merchant"],
                r["amount"],
                r["supply_amount"] if r["supply_amount"] is not None else "",
                r["vat_amount"] if r["vat_amount"] is not None else "",
                r["doc_type"],
                r["category"],
                r["biz_or_personal"],
                r["biz_reg_no"] or "",
                round(r["confidence"], 2) if r["confidence"] is not None else "",
            ]
        )
        if r["biz_or_personal"] == "사업":
            biz_total += r["amount"]
            if r["supply_amount"] is not None:
                supply_total += r["supply_amount"]
            if r["vat_amount"] is not None:
                vat_total += r["vat_amount"]

    total_row = len(rows) + 3
    ws.cell(row=total_row, column=1, value="사업 경비 합계").font = Font(bold=True)
    ws.cell(row=total_row, column=3, value=biz_total).font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=supply_total).font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=vat_total).font = Font(bold=True)

    for row_idx in range(2, total_row + 1):
        for col_idx in (3, 4, 5):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0"원"'

    for col_idx, header in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(header) + 4)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
